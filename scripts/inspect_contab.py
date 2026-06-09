# -*- coding: utf-8 -*-
import openpyxl, os

FOLDER = "//Srvad02.pedrivasco.local/publico$/VITOR CASSIMIRO/SISPRIME/CONTABILIZACOES"
names = sorted(n for n in os.listdir(FOLDER) if n.lower().endswith(".xlsx") and not n.startswith("~$"))
print("FILES:", len(names), names)

# Inspect first 3 files in detail + sheet names of all
for i, n in enumerate(names):
    wb = openpyxl.load_workbook(os.path.join(FOLDER, n), read_only=True, data_only=True)
    print(f"\n===== {n} | sheets: {wb.sheetnames}")
    if i < 4:
        for sh in wb.sheetnames:
            ws = wb[sh]
            print(f"  -- sheet '{sh}' rows={ws.max_row} cols={ws.max_column}")
            for j, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
                vals = [("" if v is None else str(v))[:20] for v in row]
                print(f"     r{j}: {vals}")
    wb.close()
