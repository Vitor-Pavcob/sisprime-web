import openpyxl, os

FOLDER = "//Srvad02.pedrivasco.local/publico$/VITOR CASSIMIRO/SISPRIME/PROPOSTAS"

print("exists:", os.path.exists(FOLDER))
names = sorted(n for n in os.listdir(FOLDER) if n.lower().endswith(".xlsx"))
print("FILES:", names)

for n in names:
    f = os.path.join(FOLDER, n)
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    resumo = next((s for s in wb.sheetnames if s.strip().lower() == "resumo"), None)
    print(f"\n===== {n} =====")
    print("  sheets:", wb.sheetnames)
    if resumo:
        ws = wb[resumo]
        print(f"  RESUMO rows={ws.max_row} cols={ws.max_column}")
        # dump first 12 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
            vals = [("" if v is None else str(v))[:22] for v in row]
            print(f"    r{i}: {vals}")
    else:
        print("  (no Resumo sheet)")
    wb.close()
