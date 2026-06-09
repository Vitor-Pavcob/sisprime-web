# -*- coding: utf-8 -*-
"""
ETL das contabilizações (baixas/recuperação) Sisprime.

Lê todos os arquivos mensais (MM.AA.xlsx) em
\\Srvad02.pedrivasco.local\publico$\VITOR CASSIMIRO\SISPRIME\CONTABILIZACOES
A competência (mês de referência) vem do NOME do arquivo (a aba interna tem
nome inconsistente — herança de template, não confiável). Ignora a aba
"Planilha1" (só legenda). Grava src/data/contabilizacoes/contabilizacoes.json + meta.

Rodar de dentro da rede:  python scripts/load-contabilizacoes.py
"""
import openpyxl, os, json, re, unicodedata, datetime

FOLDER = "//Srvad02.pedrivasco.local/publico$/VITOR CASSIMIRO/SISPRIME/CONTABILIZACOES"
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "contabilizacoes")

def norm(s):
    if s is None: return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn").strip().upper()

def fnum(v):
    if v is None: return 0.0
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if s == "": return 0.0
    if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$", s): s = s.replace(".", "").replace(",", ".")
    else: s = s.replace(",", ".")
    try: return round(float(s), 2)
    except: return 0.0

def competencia(fname):
    m = re.match(r"^(\d{2})\.(\d{2})", fname)
    return f"20{m.group(2)}-{m.group(1)}" if m else None

def tipo_norm(raw):
    n = norm(raw)
    if "ESCRIT" in n: return "Escritório 4"
    # Regra de negócio: amigável e extrajudicial são a mesma coisa → Extrajudicial.
    if "EXTRA" in n: return "Extrajudicial"
    if "AMIG" in n: return "Extrajudicial"
    if "JUDIC" in n: return "Judicial"
    return raw.strip() if raw else "Não informado"

def digits(s): return re.sub(r"\D", "", s or "")

records = []
files_info = []
for n in sorted(os.listdir(FOLDER)):
    if not n.lower().endswith(".xlsx") or n.startswith("~$"): continue
    comp = competencia(n)
    wb = openpyxl.load_workbook(os.path.join(FOLDER, n), read_only=True, data_only=True)
    data_sheet = next((s for s in wb.sheetnames if norm(s) != "PLANILHA1"), None)
    ws = wb[data_sheet]
    hdr = None; cnt = 0
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [norm(h) for h in r]; continue
        if all(c is None or str(c).strip() == "" for c in r): continue
        row = dict(zip(hdr, r))
        valor = fnum(row.get("VALOR ACORDO"))
        nome = str(row.get("NOME_CLIENTE") or "").strip()
        if valor == 0 and not nome: continue
        records.append({
            "competencia": comp,                       # 'YYYY-MM'
            "tipo": tipo_norm(row.get("ACORDO")),
            "autos": str(row.get("AUTOS") or "").strip(),
            "nome": nome,
            "cpf": str(row.get("CPF_CNPJ") or "").strip(),
            "cpfDigits": digits(row.get("CPF_CNPJ")),
            "operador": "Thayana Ramos",               # normalizado (única operadora Sisprime)
            "valor": valor,
            "ano": int(comp[:4]) if comp else None,
        })
        cnt += 1
    files_info.append({"arquivo": n, "competencia": comp, "linhas": cnt})
    wb.close()

os.makedirs(OUT_DIR, exist_ok=True)
with open(os.path.join(OUT_DIR, "contabilizacoes.json"), "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, separators=(",", ":"))

meta = {
    "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "totalContabilizacoes": len(records),
    "valorTotal": round(sum(r["valor"] for r in records), 2),
    "arquivos": files_info,
    "fonte": "\\\\Srvad02.pedrivasco.local\\publico$\\VITOR CASSIMIRO\\SISPRIME\\CONTABILIZACOES",
}
with open(os.path.join(OUT_DIR, "meta.json"), "w", encoding="utf-8") as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

print(f"OK — {len(records)} contabilizações de {len(files_info)} arquivos | R$ {meta['valorTotal']:,.2f}")
for fi in files_info:
    if fi["linhas"] == 0: print(f"  (vazio) {fi['arquivo']} [{fi['competencia']}]")
print("Gravado em:", os.path.abspath(OUT_DIR))
