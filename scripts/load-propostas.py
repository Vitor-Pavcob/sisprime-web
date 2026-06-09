# -*- coding: utf-8 -*-
"""
ETL das propostas Sisprime.

Lê SOMENTE as abas "Resumo" dos arquivos em
\\Srvad02.pedrivasco.local\publico$\VITOR CASSIMIRO\SISPRIME\PROPOSTAS
consolida, normaliza (carteira/condição/status/escritório), parseia números e
datas, e grava src/data/propostas/propostas.json + meta.json.

Local-first (igual às metas da pavcob): o JSON é bundlado no app; a página de
propostas lê o JSON, sem depender da rede em runtime.

Rodar de dentro da rede Amaral/Pavcob:  python scripts/load-propostas.py
"""
import openpyxl, os, json, unicodedata, datetime, re

FOLDER = "//Srvad02.pedrivasco.local/publico$/VITOR CASSIMIRO/SISPRIME/PROPOSTAS"
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "propostas")

def norm(s):
    if s is None: return ""
    s = "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")
    return s.strip().upper()

def fnum(v):
    if v is None: return 0.0
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if s == "": return 0.0
    # remove milhar pt-BR se vier "1.234,56"
    if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try: return round(float(s), 2)
    except: return 0.0

def to_iso(v):
    if v is None: return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m: return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{2})[/.](\d{2})[/.](\d{4})$", s)  # dd/mm/aaaa
    if m: return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None

def carteira_norm(raw):
    n = norm(raw)
    if not n or n[0].isdigit(): return "Não informada"
    # Regra de negócio: amigável e extrajudicial são a mesma coisa → Extrajudicial.
    if "EXTRA" in n: return "Extrajudicial"
    if "AMIG" in n: return "Extrajudicial"
    if "NAO AJU" in n: return "Não ajuizado"
    if any(k in n for k in ["JUDIC", "JUDC", "AJUIZ", "AJUIIZ", "AUIZ", "AZUIZ", "AJUZ"]):
        return "Judicial"
    return "Outros"

def condicao_norm(raw):
    n = norm(raw)
    if "PRAZO" in n: return "A prazo"
    if "VISTA" in n: return "À vista"
    if "DACAO" in n or "DAÇÃO" in n: return "Dação"
    return "Não informada"

def escritorio_norm(raw):
    n = norm(raw)
    if "PEDRIALI" in n: return "Pedriali & Vasconcellos"
    if "AMARAL" in n: return "Amaral Vasconcellos"
    return "—"

STATUS_MAP = {
    "PROPOSTA ACEITA": "Aceita",
    "AGUARDANDO CONTRAPROPOSTA": "Aguardando contraproposta",
    "PROPOSTA PARCELADA ENVIADA": "Parcelada enviada",
    "EM ANALISE": "Em análise",
    "BLOQUEIO JUDICIAL": "Bloqueio judicial",
}

def status_norm(raw):
    n = norm(raw)
    if not n: return None
    if "ERRO" in n or "ZIP" in n or "PERMISSION" in n or ".XLSX" in n:
        return None  # linha-lixo
    return STATUS_MAP.get(n, "Outros")

def get(row, *cands):
    for c in row:
        cn = norm(c)
        if any(norm(cand) in cn for cand in cands):
            return row[c]
    return None

def ano_arquivo(arq):
    m = re.search(r"(20\d{2})", str(arq or ""))
    return int(m.group(1)) if m else None

records = []
files_info = []
for n in sorted(os.listdir(FOLDER)):
    if not n.lower().endswith(".xlsx") or n.startswith("~$"): continue
    wb = openpyxl.load_workbook(os.path.join(FOLDER, n), read_only=True, data_only=True)
    resumo = next((s for s in wb.sheetnames if s.strip().lower() == "resumo"), None)
    if not resumo:
        wb.close(); continue
    ws = wb[resumo]
    hdr = None; cnt = 0
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [str(h).strip() if h is not None else "" for h in r]; continue
        if all(c is None or str(c).strip() == "" for c in r): continue
        row = dict(zip(hdr, r))
        status = status_norm(get(row, "Status"))
        if status is None: continue  # descarta linhas-lixo / sem status
        dataEnvio = to_iso(get(row, "Data Envio"))
        ano = int(dataEnvio[:4]) if dataEnvio else ano_arquivo(get(row, "Arquivo"))
        records.append({
            "arquivo": str(get(row, "Arquivo") or "").strip(),
            "devedor": str(get(row, "Devedor") or "").strip(),
            "cpf": str(get(row, "CPF") or "").strip(),
            "contrato": str(get(row, "Contrato") or "").strip(),
            "carteira": carteira_norm(get(row, "Carteira")),
            "carteiraRaw": str(get(row, "Carteira") or "").strip(),
            "principal": fnum(get(row, "Principal")),
            "valorAtualizado": fnum(get(row, "Valor Atualizado")),
            "valorAcordo": fnum(get(row, "Valor Acordo")),
            "entrada": fnum(get(row, "Entrada")),
            "valorParcelar": fnum(get(row, "Valor a Parcelar")),
            "parcelas": int(fnum(get(row, "Parcelas"))) if get(row, "Parcelas") else 0,
            "condicao": condicao_norm(get(row, "Condic")),
            "status": status,
            "escritorio": escritorio_norm(get(row, "Escrit")),
            "temVeiculo": bool(str(get(row, "Veiculo 1") or get(row, "Veículo 1") or "").strip()),
            "dataEnvio": dataEnvio,
            "ano": ano,
            "mes": dataEnvio[:7] if dataEnvio else None,
        })
        cnt += 1
    files_info.append({"arquivo": n, "linhas": cnt})
    wb.close()

os.makedirs(OUT_DIR, exist_ok=True)
with open(os.path.join(OUT_DIR, "propostas.json"), "w", encoding="utf-8") as f:
    json.dump(records, f, ensure_ascii=False, separators=(",", ":"))

meta = {
    "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    "totalPropostas": len(records),
    "arquivos": files_info,
    "fonte": "Resumo de \\\\Srvad02.pedrivasco.local\\publico$\\VITOR CASSIMIRO\\SISPRIME\\PROPOSTAS",
}
with open(os.path.join(OUT_DIR, "meta.json"), "w", encoding="utf-8") as f:
    json.dump(meta, f, ensure_ascii=False, indent=2)

print(f"OK — {len(records)} propostas de {len(files_info)} arquivos")
for fi in files_info: print(f"  {fi['arquivo']}: {fi['linhas']}")
print("Gravado em:", os.path.abspath(OUT_DIR))
