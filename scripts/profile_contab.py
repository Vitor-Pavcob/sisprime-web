# -*- coding: utf-8 -*-
import openpyxl, os, re, unicodedata, json
from collections import Counter

FOLDER = "//Srvad02.pedrivasco.local/publico$/VITOR CASSIMIRO/SISPRIME/CONTABILIZACOES"

def norm(s):
    if s is None: return ""
    s = "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")
    return s.strip().upper()

def fnum(v):
    if v is None: return 0.0
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if s == "": return 0.0
    if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$", s): s = s.replace(".", "").replace(",", ".")
    else: s = s.replace(",", ".")
    try: return round(float(s), 2)
    except: return 0.0

def competencia(fname):
    m = re.match(r"^(\d{2})\.(\d{2})", fname)
    if not m: return None
    return f"20{m.group(2)}-{m.group(1)}"

rows = []
for n in sorted(os.listdir(FOLDER)):
    if not n.lower().endswith(".xlsx") or n.startswith("~$"): continue
    comp = competencia(n)
    wb = openpyxl.load_workbook(os.path.join(FOLDER, n), read_only=True, data_only=True)
    data_sheet = next((s for s in wb.sheetnames if norm(s) != "PLANILHA1"), None)
    ws = wb[data_sheet]
    hdr = None
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = [norm(h) for h in r]; continue
        if all(c is None or str(c).strip() == "" for c in r): continue
        row = dict(zip(hdr, r))
        valor = fnum(row.get("VALOR ACORDO"))
        nome = str(row.get("NOME_CLIENTE") or "").strip()
        if valor == 0 and not nome: continue
        rows.append({
            "comp": comp, "file": n,
            "acordo": str(row.get("ACORDO") or "").strip(),
            "autos": str(row.get("AUTOS") or "").strip(),
            "nome": nome,
            "cpf": str(row.get("CPF_CNPJ") or "").strip(),
            "operador": str(row.get("OPERADOR") or "").strip(),
            "valor": valor,
        })
    wb.close()

print("TOTAL contabilizacoes:", len(rows))
print("VALOR total: R$ {:,.2f}".format(sum(r["valor"] for r in rows)))
print("\n=== ACORDO (tipo) ===")
for s,c in Counter(norm(r["acordo"]) for r in rows).most_common(): print(f"  {c:4} {s}")
print("\n=== OPERADOR ===")
for s,c in Counter(r["operador"] for r in rows).most_common(): print(f"  {c:4} {s}")
print("\n=== por competencia (mes) ===")
agg = {}
for r in rows:
    a = agg.setdefault(r["comp"], [0,0.0]); a[0]+=1; a[1]+=r["valor"]
for k in sorted(agg): print(f"  {k}: {agg[k][0]:3} contab | R$ {agg[k][1]:,.2f}")

# cross-ref com propostas por CPF
def digits(s): return re.sub(r"\D","", s or "")
try:
    props = json.load(open(os.path.join(os.path.dirname(__file__),"..","src","data","propostas","propostas.json"),encoding="utf-8"))
    pcpf = set(digits(p["cpf"]) for p in props if p.get("cpf"))
    ccpf = set(digits(r["cpf"]) for r in rows if r["cpf"])
    inter = {c for c in ccpf if c} & pcpf
    print(f"\n=== CROSS-REF por CPF ===")
    print(f"  CPFs distintos contab: {len([c for c in ccpf if c])}")
    print(f"  CPFs distintos propostas: {len(pcpf)}")
    print(f"  CPFs em ambos: {len(inter)}")
    matched = sum(1 for r in rows if digits(r['cpf']) in pcpf)
    print(f"  contabilizacoes com CPF em propostas: {matched}/{len(rows)}")
except Exception as e:
    print("cross-ref skip:", e)
