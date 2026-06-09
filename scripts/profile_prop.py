# -*- coding: utf-8 -*-
import openpyxl, os, unicodedata
from collections import Counter

FOLDER = "//Srvad02.pedrivasco.local/publico$/VITOR CASSIMIRO/SISPRIME/PROPOSTAS"

def norm(s):
    if s is None: return ""
    s = str(s).strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.upper()

def fnum(v):
    if v is None: return 0.0
    try: return float(str(v).replace(",", "."))
    except: return 0.0

def yr(v):
    # Data Envio -> ano
    if v is None: return None
    s = str(v)
    if len(s) >= 4 and s[:4].isdigit(): return int(s[:4])
    return None

rows = []
for n in sorted(os.listdir(FOLDER)):
    if not n.lower().endswith(".xlsx"): continue
    wb = openpyxl.load_workbook(os.path.join(FOLDER, n), read_only=True, data_only=True)
    resumo = next((s for s in wb.sheetnames if s.strip().lower() == "resumo"), None)
    if not resumo: continue
    ws = wb[resumo]
    hdr = None
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            hdr = r; continue
        if all(c is None or str(c).strip() == "" for c in r): continue
        rows.append(dict(zip([str(h).strip() for h in hdr], r)))
    wb.close()

print("TOTAL data rows:", len(rows))
cols = list(rows[0].keys())
print("COLS:", cols)

def col(row, *cands):
    for c in row:
        cn = norm(c)
        for cand in cands:
            if norm(cand) in cn: return row[c]
    return None

# Status
print("\n=== STATUS ===")
for s, c in Counter(norm(col(r,'Status')) for r in rows).most_common():
    print(f"  {c:4}  {s}")

print("\n=== CARTEIRA ===")
for s, c in Counter(norm(col(r,'Carteira')) for r in rows).most_common(30):
    print(f"  {c:4}  {s}")

print("\n=== CONDICAO ===")
for s, c in Counter(norm(col(r,'Condic')) for r in rows).most_common():
    print(f"  {c:4}  {s}")

print("\n=== ESCRITORIO ===")
for s, c in Counter(norm(col(r,'Escrit')) for r in rows).most_common():
    print(f"  {c:4}  {s}")

print("\n=== ANO (Data Envio) ===")
for s, c in sorted(Counter(yr(col(r,'Data Envio')) for r in rows).items(), key=lambda x:(x[0] is None, x[0])):
    print(f"  {s}: {c}")

print("\n=== AJUIZADO ===")
for s, c in Counter(norm(col(r,'Ajuizado')) for r in rows).most_common():
    print(f"  {c:4}  {s}")

va = sum(fnum(col(r,'Valor Acordo')) for r in rows)
vat = sum(fnum(col(r,'Valor Atualizado')) for r in rows)
vp = sum(fnum(col(r,'Principal')) for r in rows)
print("\n=== SOMAS ===")
print(f"  Valor Acordo total:     {va:,.2f}")
print(f"  Valor Atualizado total: {vat:,.2f}")
print(f"  Principal total:        {vp:,.2f}")
print(f"  Desagio medio (1 - acordo/atualizado): {(1 - va/vat)*100:.1f}%")
nveic = sum(1 for r in rows if (col(r,'Veiculo 1') or '').strip())
print(f"  Propostas com veiculo:  {nveic}")
print(f"  CPFs distintos:         {len(set(norm(col(r,'CPF')) for r in rows if col(r,'CPF')))}")
